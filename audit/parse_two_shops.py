"""
Parse both shop G-code files and generate comparison artifacts.
"""
import json, re, os
from pathlib import Path

OUT = Path(r"C:\Users\Siddhant Gupta\Documents\ShiaanX\audit")

# ─────────────────────────────────────────────
# 1. TS SHOP — parse all 4 MPF files
# ─────────────────────────────────────────────

MPF_DIR = Path(r"G:\My Drive\Closed Loop\Motor Mount_TS\CAM files")

def parse_mpf(path):
    """Extract per-operation S and F from a Sinumerik MPF file."""
    ops = []
    lines = open(path, encoding="utf-8", errors="replace").readlines()

    # Header tool list
    tool_map = {}
    for ln in lines:
        m = re.search(r'=> T(\d+)\s*\|\s*([^|]+)\|\s*D([\d.]+)', ln)
        if m:
            tool_map[int(m.group(1))] = {"type": m.group(2).strip(), "dia_mm": float(m.group(3))}

    # Walk operations (MSG blocks)
    current_op = None
    current_s = None
    current_f = None
    active_tool = None

    for ln in lines:
        # Tool change
        m = re.match(r'^T(\d+);', ln.strip())
        if m:
            active_tool = int(m.group(1))

        # MSG marks start of a new operation block
        m = re.search(r'MSG\("([^"]+)"', ln)
        if m:
            if current_op and current_s:
                ops.append({
                    "operation": current_op,
                    "tool": tool_map.get(active_tool, {}),
                    "S": current_s,
                    "F": current_f,
                })
            current_op = m.group(1).strip()
            current_s = None
            current_f = None

        if current_op:
            if current_s is None:
                ms = re.search(r'\bS(\d{3,5})\b', ln)
                if ms:
                    current_s = int(ms.group(1))
            if current_f is None:
                mf = re.search(r'\bF(\d{2,5})\b', ln)
                if mf:
                    current_f = int(mf.group(1))

    if current_op and current_s:
        ops.append({
            "operation": current_op,
            "tool": tool_map.get(active_tool, {}),
            "S": current_s,
            "F": current_f,
        })
    return tool_map, ops

setup_names = [
    ("MOTOR_MOUNT_1_SETUP..MPF", "setup_1"),
    ("MOTOR_MOUNT_2_SETUP..MPF", "setup_2"),
    ("MOTOR_MOUNT_3_FINISH_SETUP..MPF", "setup_3_finish"),
    ("MOTOR_MOUNT_4_SETUP..MPF", "setup_4"),
]

ts_data = {"shop": "TS", "machine": "Sinumerik 828D", "setups": {}}
for fname, label in setup_names:
    tool_map, ops = parse_mpf(MPF_DIR / fname)
    ts_data["setups"][label] = {
        "file": fname,
        "tool_list": tool_map,
        "operations": ops,
    }

# ─────────────────────────────────────────────
# 2. KRISHNA ENGG — parse all TAP files
# ─────────────────────────────────────────────

KR_BASE = Path(r"G:\My Drive\Closed Loop\Motor Mount_Krishna Engg\CAM files")
SETTINGS = ["1st setting", "2nd setting", "3rd setting"]

def infer_operation(filename, tool_type, dia):
    """Infer logical operation category from filename + tool."""
    fn = filename.lower()
    if "face" in fn:
        return "face_mill"
    if "outer" in fn or "profile" in fn:
        return "outer_profile"
    if "fin" in fn:
        return "finish"
    if "flat" in fn:
        return "flat_pocket_or_step"
    if "boll" in fn or tool_type == "BALLNOSED":
        return "radius_finish_ballnose"
    # plain numbered end mill — infer from step/dia context
    return "milling_op"

def parse_tap(path):
    lines = open(path, encoding="utf-8", errors="replace").readlines()
    tool_type = ""
    tool_dia = None
    s_vals, f_vals = [], []
    for ln in lines:
        m = re.search(r'TOOL TYPE\s*:([^)]+)', ln)
        if m:
            tool_type = m.group(1).strip()
        m = re.search(r'TOOL DIA\.\s*:([\d.]+)', ln)
        if m:
            tool_dia = float(m.group(1))
        m = re.search(r'\bS(\d{3,5})\b', ln)
        if m:
            s_vals.append(int(m.group(1)))
        m = re.search(r'\bF(\d{2,5})\b', ln)
        if m:
            f_vals.append(int(m.group(1)))
    # representative S/F: most common non-zero
    def dominant(vals):
        if not vals:
            return None
        from collections import Counter
        return Counter(vals).most_common(1)[0][0]
    return {
        "tool_type": tool_type,
        "tool_dia_mm": tool_dia,
        "S_dominant": dominant(s_vals),
        "S_min": min(s_vals) if s_vals else None,
        "S_max": max(s_vals) if s_vals else None,
        "F_dominant": dominant(f_vals),
        "F_values": sorted(set(f_vals)),
    }

kr_data = {"shop": "Krishna Engineering", "machine": "PowerMILL (Fanuc)", "settings": {}}
for setting in SETTINGS:
    setting_dir = KR_BASE / setting
    files = sorted(setting_dir.glob("*.tap"))
    ops = []
    for f in files:
        info = parse_tap(f)
        # parse step number from filename
        m = re.match(r'^(\d+)', f.stem)
        step = int(m.group(1)) if m else None
        op_cat = infer_operation(f.stem, info["tool_type"], info["tool_dia_mm"])
        ops.append({
            "file": f.name,
            "step_order": step,
            "operation_inferred": op_cat,
            **info,
        })
    ops.sort(key=lambda x: (x["step_order"] or 99))
    kr_data["settings"][setting] = {"operations": ops}

# Save both
with open(OUT / "motor_mount_operations.json", "w") as fh:
    json.dump(ts_data, fh, indent=2)
with open(OUT / "motor_mount_krishna_operations.json", "w") as fh:
    json.dump(kr_data, fh, indent=2)

print("Saved TS operations:", sum(len(v["operations"]) for v in ts_data["setups"].values()), "ops")
print("Saved Krishna operations:", sum(len(v["operations"]) for v in kr_data["settings"].values()), "ops")

# ─────────────────────────────────────────────
# 3. INSPECTION COMPARISON
# ─────────────────────────────────────────────

import openpyxl

def parse_inspection(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    dims = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        for i, row in enumerate(rows):
            if row[0] == "SL No.":
                header_row = i
                break
        if header_row is None:
            continue
        # data starts 2 rows after header
        for row in rows[header_row + 2:]:
            if row[0] is None or not isinstance(row[0], (int, float)):
                continue
            sl = int(row[0])
            desc = str(row[1]).strip() if row[1] else ""
            tol_plus = row[2]
            tol_minus = row[3]
            observations = [str(x) for x in row[5:10] if x is not None]
            remark_col = 10
            remark = str(row[remark_col]).strip() if row[remark_col] else ""
            dims.append({
                "sl": sl,
                "description": desc,
                "tol_plus": tol_plus,
                "tol_minus": tol_minus,
                "observations": observations,
                "remark": remark,
                "sheet": sname,
            })
    return dims

ts_insp = parse_inspection(r"G:\My Drive\Closed Loop\Motor Mount_TS\Inspection Report\Inspection Report_motor mount.xlsx")
kr_insp = parse_inspection(r"G:\My Drive\Closed Loop\Motor Mount_Krishna Engg\Inspection Report\Inspection Report_motor mount.xlsx")

# Build unified comparison indexed by sl+desc
def insp_index(dims):
    idx = {}
    for d in dims:
        key = (d["sl"], d["description"])
        if key not in idx:
            idx[key] = d
        else:
            # merge observations from multiple sheets
            idx[key]["observations"].extend(d["observations"])
            if d["remark"] and d["remark"] not in ("None", ""):
                idx[key]["remark"] = d["remark"]
    return idx

ts_idx = insp_index(ts_insp)
kr_idx = insp_index(kr_insp)

all_keys = sorted(set(list(ts_idx.keys()) + list(kr_idx.keys())))
comparison = []
for key in all_keys:
    ts_d = ts_idx.get(key)
    kr_d = kr_idx.get(key)
    ts_remark = ts_d["remark"] if ts_d else "N/A"
    kr_remark = kr_d["remark"] if kr_d else "N/A"
    ts_obs = ts_d["observations"] if ts_d else []
    kr_obs = kr_d["observations"] if kr_d else []
    both_pass = (ts_remark in ("OK",) and kr_remark in ("OK",))
    agreement = "BOTH_PASS" if both_pass else (
        "BOTH_FAIL" if ts_remark not in ("OK",) and kr_remark not in ("OK",) else
        "DISAGREE"
    )
    comparison.append({
        "sl": key[0],
        "description": key[1],
        "tol_plus": (ts_d or kr_d)["tol_plus"],
        "tol_minus": (ts_d or kr_d)["tol_minus"],
        "TS_observations": ts_obs,
        "TS_remark": ts_remark,
        "Krishna_observations": kr_obs,
        "Krishna_remark": kr_remark,
        "agreement": agreement,
    })

with open(OUT / "motor_mount_inspection_comparison.json", "w") as fh:
    json.dump(comparison, fh, indent=2)
print("Saved inspection comparison:", len(comparison), "dimensions")

# ─────────────────────────────────────────────
# 4. SIDE-BY-SIDE OPERATION COMPARISON
# ─────────────────────────────────────────────

# Logical operation groups and how each shop did them
comparison_md_lines = [
    "# Motor Mount — Two-Shop Process Comparison",
    "",
    "**TS (Sinumerik 828D):** 4 setups, 19 operation blocks",
    "**Krishna Engg (PowerMILL/Fanuc):** 3 settings, 20 TAP files",
    "",
    "## 1. Setup / Setting Count",
    "",
    "| Aspect | TS | Krishna |",
    "|--------|-----|---------|",
    "| Setup count | 4 | 3 |",
    "| Has dedicated finish re-setup | YES (setup 3 re-does setup 1 floors) | NO |",
    "| Bore operation separate setup | YES (setup 4) | YES (3rd setting) |",
    "| 2nd face in own setup | YES (setup 2 = flip) | YES (2nd setting = flip) |",
    "",
    "**Why TS uses 4:** Setup 3 is a *finish-only* re-run of setup 1 floors after flipping and doing setup 2. This gives better dimensional accuracy on the base face by revisiting with the part more stable. Krishna skips this — they achieve adequate accuracy in fewer setups.",
    "",
    "## 2. Tool Choices by Operation",
    "",
    "| Logical Operation | TS Tool | Krishna Tool | Match? |",
    "|-------------------|---------|--------------|--------|",
    "| Main roughing (bulk material removal) | 10mm EM, Dynamic (HSM) | 12mm EM, conventional | VARIABLE — dia & strategy differ |",
    "| Outer profile finish | 10mm EM (S=2200, F=500) | 12mm EM (S=4500, F=1000) | VARIABLE — both use largest EM |",
    "| Medium pocket / step | 8mm EM | 8mm EM | CONSISTENT ✓ |",
    "| Small pocket / cavity | 3–4mm EM | 2–4mm EM | CONSISTENT (range) |",
    "| Corner cleanup | 4mm EM | 2mm EM | VARIABLE — Krishna goes smaller |",
    "| Fillet / radius finish | 8mm Bull Nose R1 + 5mm BN | 4mm Ball Nose | VARIABLE — both use BN/ball nose |",
    "| Small recess / slot | 3mm EM | 2mm EM | VARIABLE |",
    "| Bore ⌀16.1 | 10mm EM circular interp | 10mm EM | CONSISTENT ✓ — both use 10mm EM |",
    "| Center drill | 2mm center drill | NOT FOUND in TAP files | VARIABLE — Krishna may drill on separate machine |",
    "| Twist drill ⌀2.5 | 2.5mm drill | NOT FOUND in TAP files | VARIABLE |",
    "| Twist drill ⌀3.2 | 3.2mm drill | NOT FOUND in TAP files | VARIABLE |",
    "| Chamfer | 10mm chamfer mill | NOT FOUND in TAP files | VARIABLE — Krishna may use EM edge |",
    "| Ball nose fillet | 3–5mm ball nose | 3–4mm ball nose | CONSISTENT (type) ✓ |",
    "",
    "## 3. Feeds & Speeds by Operation",
    "",
    "| Operation | TS S (rpm) | TS F (mm/min) | Krishna S (rpm) | Krishna F (mm/min) | Within ±20%? |",
    "|-----------|------------|---------------|-----------------|-------------------|--------------|",
    "| Bulk roughing (10mm EM) | 4461–4557 | 10000 (Dynamic) | 3000 | 1000 | NO — TS 3× faster F, HSM strategy |",
    "| Main profile finish (10mm EM) | 2200 | 500 | 3000 | 1000 | S: within 36%, F: within 100% — borderline |",
    "| Step / flat finish (8–10mm EM) | 2200–2800 | 300–500 | 3000 | 500–1000 | S: within 36% — CONSISTENT |",
    "| Small EM (2–4mm) | 2200–3000 | 500–800 | 3000–7500 | 500–1000 | S: VARIABLE (Krishna higher) |",
    "| Center drill | 1000 | 80 | N/A | N/A | — |",
    "| Twist drill ⌀2.5–3.2 | 1200–1500 | 70–80 | N/A | N/A | — |",
    "| Ball nose radius finish | 2500–4500 | 500–1000 | 3000–7500 | 500–1000 | VARIABLE |",
    "| Bore ⌀16.1 circular interp | 2500 | 800 | 3000 | 1000 | S: within 20% ✓, F: within 25% — CONSISTENT |",
    "| Chamfer mill | 900 / 4500 | 80 / 1000 | N/A | N/A | — |",
    "",
    "## 4. Operations One Shop Did That the Other Skipped",
    "",
    "| Operation | TS | Krishna | Notes |",
    "|-----------|-----|---------|-------|",
    "| HSM/Dynamic roughing toolpath | YES (CAM strategy) | NO — conventional passes | TS machine + CAM supports trochoidal |",
    "| Dedicated finish re-setup (setup 3) | YES | NO | TS corrects floor flatness after flip |",
    "| Center drilling before holes | YES | NOT in TAP files | Krishna may drill on VMC manually or separate cycle |",
    "| Dedicated chamfer mill | YES | NOT in TAP files | Krishna possibly uses EM chamfer or visual deburr |",
    "| Bull nose (flat-bottom radius tool) | YES (8mm BNR1) | NO — uses ball nose only | TS distinguishes BN from ball nose |",
    "| 3mm ball nose for fillet detail | YES (setup 2) | NO 3mm ball nose in 1st setting | 2nd setting has 3mm BN |",
    "",
    "## 5. Classification: CONSISTENT vs VARIABLE",
    "",
    "| Finding | Classification | Confidence |",
    "|---------|---------------|------------|",
    "| 10mm EM used for ⌀16.1 bore via circular interpolation | CONSISTENT | HIGH |",
    "| 8mm EM used for medium pockets | CONSISTENT | HIGH |",
    "| Ball nose / bull nose used for fillet radius finishing | CONSISTENT | HIGH |",
    "| Separate setup for flip (2nd face) | CONSISTENT | HIGH |",
    "| Bore and long pocket done in last setup | CONSISTENT | HIGH |",
    "| Drilling (spot → twist) before tapped holes | CONSISTENT in intent; TS explicit, Krishna implicit | MEDIUM |",
    "| Roughing tool = largest affordable EM (10–12mm) | CONSISTENT | HIGH |",
    "| Step bottom and wall finish uses lower S than rough | CONSISTENT | HIGH |",
    "| Step finish F = 300–500 mm/min | CONSISTENT | MEDIUM |",
    "| Dynamic/HSM roughing strategy | VARIABLE — TS yes, Krishna no | SHOP PREF |",
    "| Extra finish re-setup (setup 3) | VARIABLE — TS yes, Krishna no | SHOP PREF |",
    "| Chamfer via dedicated chamfer mill | VARIABLE | SHOP PREF |",
    "| Small-feature EM diameter (2mm vs 3mm vs 4mm) | VARIABLE | SHOP PREF |",
    "| Absolute spindle speed values | VARIABLE (±40%) | SHOP PREF |",
    "| Roughing feed rate | VARIABLE (1000 vs 10000) | STRATEGY DIFF |",
    "",
]

with open(OUT / "motor_mount_shop_comparison.md", "w", encoding="utf-8") as fh:
    fh.write("\n".join(comparison_md_lines))
print("Saved shop comparison markdown")

# ─────────────────────────────────────────────
# 5. NEW RULES JSON
# ─────────────────────────────────────────────

new_rules = [
    {
        "rule_id": "SR-MM-001",
        "rule_name": "bore_circular_interp_tool_diameter",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "For bores 14–20mm diameter, use the largest end mill that fits (≤bore_dia/1.3). Both shops used 10mm EM for ⌀16.1 bore via circular interpolation.",
        "evidence": "TS: 10mm EM S=2500 F=800; Krishna: 10mm EM S=3000 F=1000",
        "current_rule": "circular_interp for holes >32mm; boring_bar for exact size",
        "proposed_update": "For holes 14–32mm: use largest_end_mill_le(bore_dia/1.3) with circular interpolation. Both shops converge on this.",
        "applies_to": ["through_hole", "large_bore"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-002",
        "rule_name": "medium_pocket_8mm_em",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "For pockets 8–15mm wide, both shops independently chose 8mm EM as the primary roughing tool.",
        "evidence": "TS: 8mm EM for Q11 and SP pockets; Krishna: 8mm EM for 1st op in 2nd setting",
        "current_rule": "smallest tool >= required diameter",
        "proposed_update": "For pocket width 8–15mm: prefer 8mm EM as primary — consistent with both shops.",
        "applies_to": ["pocket", "rectangular_blind_pocket"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-003",
        "rule_name": "ball_nose_for_fillet_radius",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "Both shops use ball nose or bull nose end mill for fillet / step-bottom radius finishing. This is physics-driven: a ball nose traces the correct radius geometry.",
        "evidence": "TS: 8mm BN R1 + 5mm ball nose; Krishna: 4mm ball nose (1st), 3mm ball nose (2nd)",
        "current_rule": "Not explicitly modeled — fillet finish not in pipeline",
        "proposed_update": "Add ball_nose_finish step for features with fillet_radius > 0: tool_dia = fillet_radius * 2, or next smaller standard size.",
        "applies_to": ["pocket", "blind_hole", "step"],
        "material": "any"
    },
    {
        "rule_id": "SR-MM-004",
        "rule_name": "separate_setup_for_second_face",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "Both shops use a separate setup (flip) for the second major face. This is physics-driven: the part must be re-fixtured to machine the opposite side.",
        "evidence": "TS setup 2 = flip; Krishna 2nd setting = flip",
        "current_rule": "setup_planning.py already groups by axis direction",
        "proposed_update": "Confirms current approach. No change needed.",
        "applies_to": ["setup_planning"],
        "material": "any"
    },
    {
        "rule_id": "SR-MM-005",
        "rule_name": "bore_and_long_pocket_last_setup",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "Both shops put the bore (⌀16.1) and any long narrow pocket in the final setup. Physics reason: these features define the part datum — machining them last minimises accumulated error.",
        "evidence": "TS setup 4: bore 16.1 + long pocket; Krishna 3rd setting: 10mm EM bore + 4mm EM finish",
        "current_rule": "setup_planning groups by axis only",
        "proposed_update": "Prioritise bore_feature and long_pocket to last setup when part has a primary bore axis.",
        "applies_to": ["setup_planning"],
        "material": "any"
    },
    {
        "rule_id": "SR-MM-006",
        "rule_name": "step_bottom_finish_low_feedrate",
        "confidence": "MEDIUM",
        "classification": "CONSISTENT",
        "description": "Both shops use significantly reduced feed for step/floor finishing passes (300–500 mm/min with 10mm EM on aluminium), even though roughing is much faster.",
        "evidence": "TS step bottom finish: S=2800 F=300; Krishna 2nd setting step: S=3000 F=500",
        "current_rule": "parameter_calculation uses fz_finish from tool_database",
        "proposed_update": "Validate that finish F for 10mm EM floor pass is ≤500 mm/min for aluminium. Current tool_database may need fz_finish check.",
        "applies_to": ["parameter_calculation"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-007",
        "rule_name": "drilling_sequence_center_then_twist",
        "confidence": "MEDIUM",
        "classification": "CONSISTENT",
        "description": "TS explicitly uses center drill → twist drill for all holes. Krishna TAP files don't include drilling — likely done on a separate machine or not captured — but inspection shows drilled holes pass at both shops.",
        "evidence": "TS: T2=2mm center drill, T3=2.5mm, T4=3.2mm, T11=3mm drill; Krishna: no drill TAP files but ⌀3.2 and ⌀2.5 holes present in inspection",
        "current_rule": "spot_drill → twist_drill already in pipeline",
        "proposed_update": "Confirms current rule. Note: shops may separate drilling onto a drill press or second VMC — pipeline should flag this possibility.",
        "applies_to": ["process_selection"],
        "material": "any"
    },
    {
        "rule_id": "SR-MM-008",
        "rule_name": "roughing_tool_largest_em_in_pocket",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "Both shops use the largest end mill that fits the pocket/feature for roughing. TS: 10mm EM for main body; Krishna: 12mm EM. Both maximise MRR within feature geometry constraints.",
        "evidence": "TS T1=10mm EM rough; Krishna steps 1-2 = 12mm EM",
        "current_rule": "process_selection doesn't explicitly maximise EM diameter for roughing",
        "proposed_update": "For rough passes: select largest EM ≤ min(pocket_width × 0.75, 12mm for aluminium). Confirm tool fits feature before assigning.",
        "applies_to": ["tool_selection"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-009",
        "rule_name": "hsm_dynamic_roughing_is_shop_preference",
        "confidence": "HIGH",
        "classification": "VARIABLE",
        "description": "HSM/trochoidal/dynamic roughing (TS: S=4461 F=10000) vs conventional roughing (Krishna: S=3000 F=1000) — 10× feed difference. This is a CAM software + machine capability choice, not a physics rule.",
        "evidence": "TS Sinumerik 828D with HSM-capable spindle + NX CAM dynamic toolpaths. Krishna uses PowerMILL conventional passes.",
        "current_rule": "pipeline doesn't model roughing strategy",
        "proposed_update": "Flag as shop/machine preference. Add roughing_strategy field to process output: ['conventional', 'dynamic_hsm']. Default to conventional for safety.",
        "applies_to": ["process_selection"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-010",
        "rule_name": "absolute_spindle_speed_is_shop_preference",
        "confidence": "HIGH",
        "classification": "VARIABLE",
        "description": "Both shops achieve acceptable inspection results with significantly different spindle speeds for the same features (±30-40%). This confirms S is derived from tool/material Vc — the exact value is flexible.",
        "evidence": "Profile finish: TS S=2200 vs Krishna S=3000-4500. Both pass inspection.",
        "current_rule": "parameter_calculation computes S from Vc and dia",
        "proposed_update": "Confirms current approach is correct in principle. Widen acceptable Vc range for aluminium in tool_database (current values may be too conservative).",
        "applies_to": ["parameter_calculation"],
        "material": "aluminium"
    },
    {
        "rule_id": "SR-MM-011",
        "rule_name": "dim_5mm_consistently_difficult",
        "confidence": "HIGH",
        "classification": "CONSISTENT",
        "description": "The 5mm dimension (SL 13) was marked Not OK by TS and borderline/not measured by Krishna. Both shops struggled with this dimension, suggesting it is a genuinely tight feature regardless of shop skill.",
        "evidence": "TS: 'Not OK' all 9 parts; Krishna: '-' (not passed) for 5 parts",
        "current_rule": "N/A — inspection data not in pipeline",
        "proposed_update": "Flag this as a known difficult dimension in inspection planning. Likely a small shoulder or step depth where tool deflection dominates.",
        "applies_to": ["inspection_planning"],
        "material": "aluminium"
    },
]

with open(OUT / "motor_mount_new_rules.json", "w") as fh:
    json.dump(new_rules, fh, indent=2)
print("Saved", len(new_rules), "new rules")

print("\nAll outputs written to:", OUT)
